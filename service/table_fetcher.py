import json
import io
import os
import hashlib
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
from openpyxl import load_workbook

CONFIG_DIR = '../config'

def file_hash(file_path):
    hash_sha256 = hashlib.sha256()
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_sha256.update(chunk)
    return hash_sha256.hexdigest()

def update_data_json():
    wb = load_workbook(filename='data/groups.xlsx')

    with open(f'{CONFIG_DIR}/config.json', 'r') as file:
        sheet_name = json.load(file)['sheet_name']

    sheet = wb[sheet_name]

    data = {}
    for row in sheet.rows:
        data[row[2].value] = [i.value for i in row[4:sheet.max_column-1] if i.value is not None]

    del data[sheet['C1'].value]
    

    with open('data/groups.json', 'w') as file:
        json.dump(data, file)

def main():
    with open('../config/client.json', 'r') as file:
        api_key = json.load(file)['web']['api_key']

    client = build('drive', 'v3', developerKey=api_key)

    with open(f'{CONFIG_DIR}/config.json', 'r') as file:
        file_id = json.load(file)['groups_sheet']

    request = client.files().get_media(fileId=file_id)

    new = io.BytesIO()
    downloader = MediaIoBaseDownload(new, request)
    done = False
    while done is False:
        _, done = downloader.next_chunk()

    new_file_path = 'data/temp_groups.xlsx'
    with open(new_file_path, 'wb') as f:
        f.write(new.getvalue())

    existing_file_path = 'data/groups.xlsx'

    if os.path.exists(existing_file_path):
        if file_hash(existing_file_path) == file_hash(new_file_path):
            print("Files are the same. No action needed.")
        else:
            os.rename(new_file_path, existing_file_path)
            print("Files are different. Updated the file.")
            update_data_json()
    else:
        os.rename(new_file_path, existing_file_path)
        print("File does not exist. Created new file.")
        update_data_json()

    if os.path.exists(new_file_path):
        os.remove(new_file_path)
if __name__ == '__main__':
    main()


    